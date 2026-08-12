from pathlib import Path
from shutil import copy2
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

ROOT = Path(".")
REPORT = ROOT / "reports" / "EBRIP_Executive_Risk_Report.xlsx"
BACKUP = ROOT / "reports" / "EBRIP_Executive_Risk_Report_before_dashboard.xlsx"

# Safety backup
copy2(REPORT, BACKUP)

wb = load_workbook(REPORT)

# ---------- Theme ----------
NAVY = "1F4E78"
BLUE = "5B9BD5"
LIGHT_BLUE = "D9EAF7"
GREEN = "70AD47"
LIGHT_GREEN = "E2F0D9"
ORANGE = "ED7D31"
LIGHT_ORANGE = "FCE4D6"
RED = "C00000"
LIGHT_RED = "F4CCCC"
GRAY = "F2F2F2"
DARK = "404040"
WHITE = "FFFFFF"

thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_range(ws, cell_range, fill=None, font=None, alignment=None):
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            cell.border = border

def clear_dashboard(ws):
    # Remove existing charts and images.
    ws._charts = []
    ws._images = []

    # Remove merged cells first so MergedCell objects become normal cells.
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))

    # Clear existing values and styles.
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
            cell._style = cell._style

def add_kpi(ws, cell, title, value, fill_color):
    col = ws[cell].column
    row = ws[cell].row

    title_cell = ws.cell(row, col, title)
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.font = Font(color=WHITE, bold=True, size=10)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = border

    value_cell = ws.cell(row + 1, col, value)
    value_cell.fill = PatternFill("solid", fgColor=fill_color)
    value_cell.font = Font(color=DARK, bold=True, size=16)
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.border = border

def set_widths(ws):
    widths = {
        "A": 24, "B": 22, "C": 18, "D": 18,
        "E": 24, "F": 18, "G": 18, "H": 18,
        "I": 18, "J": 18, "K": 18, "L": 18
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

# ---------- Read existing model data ----------
integrated = wb["Integrated Risk"]
credit = wb["Credit Risk"]
fraud = wb["Fraud Detection"]
txn = wb["Transaction Fraud"]
aml = wb["AML Monitoring"]

# Known validated values from existing workbook
overall_score = 0.0
risk_tier = "LOW"
alert_status = "NO_ALERT"

credit_prauc = credit["G4"].value
fraud_prauc = fraud["G6"].value
txn_prauc = txn["G5"].value
aml_rocauc = aml["B11"].value

high_risk = credit["B12"].value
very_high_risk = credit["B13"].value

# ---------- Executive Summary ----------
ws = wb["Executive Summary"]
clear_dashboard(ws)

ws.merge_cells("A1:L2")
ws["A1"] = "Enterprise Banking Risk Intelligence Platform"
ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
ws["A1"].font = Font(color=WHITE, bold=True, size=20)
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

ws.merge_cells("A3:L3")
ws["A3"] = "Executive Risk Dashboard | Credit • Fraud • Transaction Fraud • AML"
ws["A3"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws["A3"].font = Font(color=NAVY, bold=True, size=11)
ws["A3"].alignment = Alignment(horizontal="left", vertical="center")

# KPI cards
add_kpi(ws, "A5", "OVERALL RISK TIER", risk_tier, LIGHT_GREEN)
add_kpi(ws, "C5", "OVERALL RISK SCORE", overall_score, LIGHT_GREEN)
add_kpi(ws, "E5", "ALERT STATUS", alert_status, LIGHT_GREEN)
add_kpi(ws, "G5", "HIGH RISK", high_risk, LIGHT_ORANGE)
add_kpi(ws, "I5", "VERY HIGH RISK", very_high_risk, LIGHT_RED)
add_kpi(ws, "K5", "TXN FRAUD PR-AUC", round(txn_prauc, 4), LIGHT_GREEN)

for col in range(1, 13):
    ws.column_dimensions[get_column_letter(col)].width = 18

ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 10
ws.row_dimensions[5].height = 24
ws.row_dimensions[6].height = 30

# Domain performance table
ws["A9"] = "Cross-Domain Model Performance"
ws["A9"].fill = PatternFill("solid", fgColor=NAVY)
ws["A9"].font = Font(color=WHITE, bold=True, size=12)
ws.merge_cells("A9:F9")

headers = ["Risk Domain", "Selected Model", "ROC-AUC", "PR-AUC"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(10, c, h)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.font = Font(color=WHITE, bold=True)
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

domain_rows = [
    ["Credit Risk", "Gradient Boosting", 0.871633, 0.409072],
    ["Fraud Detection", "Gradient Boosting", 0.868241, 0.407716],
    ["Transaction Fraud", "Random Forest", 0.972772, 0.807150],
    ["AML Monitoring", "Logistic Regression", 0.861643, 0.000361],
]

for r, row in enumerate(domain_rows, 11):
    for c, value in enumerate(row, 1):
        cell = ws.cell(r, c, value)
        cell.border = border
        if c >= 3:
            cell.number_format = "0.0000"

# Governance warning
ws.merge_cells("G9:L9")
ws["G9"] = "MODEL GOVERNANCE"
ws["G9"].fill = PatternFill("solid", fgColor=ORANGE)
ws["G9"].font = Font(color=WHITE, bold=True, size=12)
ws["G9"].alignment = Alignment(horizontal="left")

ws.merge_cells("G10:L14")
ws["G10"] = (
    "AML validation is currently experimental. "
    "The validation dataset contains only 1 positive AML case out of 20,000 rows. "
    "Threshold metrics should not be treated as production-grade until evaluated "
    "on a larger and more representative AML dataset."
)
ws["G10"].fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
ws["G10"].font = Font(color=DARK, bold=True, size=11)
ws["G10"].alignment = Alignment(wrap_text=True, vertical="top")
style_range(ws, "G10:L14")

# Risk summary
ws["A17"] = "Business Risk Interpretation"
ws["A17"].fill = PatternFill("solid", fgColor=NAVY)
ws["A17"].font = Font(color=WHITE, bold=True, size=12)
ws.merge_cells("A17:L17")

ws.merge_cells("A18:L20")
ws["A18"] = (
    "The integrated demonstration engine currently classifies the assessed entity as LOW risk "
    "with NO_ALERT status. Credit and fraud models use business-oriented thresholds designed "
    "to improve detection recall. Transaction fraud provides the strongest model discrimination "
    "in the current validation results, while AML requires additional representative data."
)
ws["A18"].alignment = Alignment(wrap_text=True, vertical="top")
ws["A18"].fill = PatternFill("solid", fgColor=GRAY)
style_range(ws, "A18:L20")

# ---------- Chart 1: PR-AUC ----------
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "PR-AUC by Risk Domain"
chart.y_axis.title = "PR-AUC"
chart.x_axis.title = "Risk Domain"
data = Reference(ws, min_col=4, min_row=10, max_row=14)
cats = Reference(ws, min_col=1, min_row=11, max_row=14)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 7
chart.width = 12
chart.legend = None
ws.add_chart(chart, "A22")

# ---------- Chart 2: ROC-AUC ----------
chart2 = BarChart()
chart2.type = "col"
chart2.style = 11
chart2.title = "ROC-AUC by Risk Domain"
chart2.y_axis.title = "ROC-AUC"
chart2.x_axis.title = "Risk Domain"
data2 = Reference(ws, min_col=3, min_row=10, max_row=14)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats)
chart2.height = 7
chart2.width = 12
chart2.legend = None
ws.add_chart(chart2, "G22")

# ---------- Chart 3: Credit Risk Distribution ----------
ws["A38"] = "Credit Risk Segment Distribution"
ws["A38"].fill = PatternFill("solid", fgColor=NAVY)
ws["A38"].font = Font(color=WHITE, bold=True, size=12)
ws.merge_cells("A38:D38")

risk_data = [
    ("Low Risk", 27153),
    ("Moderate Risk", 677),
    ("High Risk", 1522),
    ("Very High Risk", 646),
]

ws["A39"] = "Risk Segment"
ws["B39"] = "Count"
for c in ["A39", "B39"]:
    ws[c].fill = PatternFill("solid", fgColor=BLUE)
    ws[c].font = Font(color=WHITE, bold=True)
    ws[c].border = border

for r, (segment, count) in enumerate(risk_data, 40):
    ws.cell(r, 1, segment)
    ws.cell(r, 2, count)
    ws.cell(r, 1).border = border
    ws.cell(r, 2).border = border

dchart = DoughnutChart()
dchart.title = "Credit Risk Distribution"
data3 = Reference(ws, min_col=2, min_row=39, max_row=43)
labels3 = Reference(ws, min_col=1, min_row=40, max_row=43)
dchart.add_data(data3, titles_from_data=True)
dchart.set_categories(labels3)
dchart.height = 8
dchart.width = 11
dchart.dataLabels = DataLabelList()
dchart.dataLabels.showPercent = True
ws.add_chart(dchart, "D39")

# ---------- Final formatting ----------
ws.freeze_panes = "A5"
ws.sheet_view.showGridLines = False

# Make all other sheets cleaner
for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = False
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal or "left",
                    vertical="center",
                    wrap_text=True
                )

# Save
wb.save(REPORT)

print("DASHBOARD GENERATED SUCCESSFULLY")
print(f"File: {REPORT.resolve()}")
print(f"Backup: {BACKUP.resolve()}")
print("Charts added: 3")
print("KPI cards added: 6")
print("Governance warning added: AML")



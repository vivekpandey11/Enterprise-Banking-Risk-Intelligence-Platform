from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import json
import csv

ROOT = Path(".")
OUTPUT = ROOT / "reports" / "EBRIP_Executive_Risk_Report.xlsx"

wb = load_workbook(OUTPUT)

# ---------- Styles ----------
title_fill = PatternFill("solid", fgColor="1F4E78")
section_fill = PatternFill("solid", fgColor="D9EAF7")
header_fill = PatternFill("solid", fgColor="5B9BD5")
white_font = Font(color="FFFFFF", bold=True, size=14)
header_font = Font(color="FFFFFF", bold=True)
bold_font = Font(bold=True)

thin = Side(style="thin", color="B7B7B7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def clear_sheet(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
            cell._style = cell._style.copy()

def write_title(ws, title):
    ws["A1"] = title
    ws["A1"].fill = title_fill
    ws["A1"].font = white_font
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.merge_cells("A1:F1")

TABLE_COUNTER = 0

TABLE_COUNTER = 0

def write_table(ws, start_row, headers, rows):
    global TABLE_COUNTER

    for col, header in enumerate(headers, 1):
        c = ws.cell(start_row, col, header)
        c.fill = header_fill
        c.font = header_font
        c.border = border
        c.alignment = Alignment(horizontal="center")

    for r, row in enumerate(rows, start_row + 1):
        for cidx, value in enumerate(row, 1):
            cell = ws.cell(r, cidx, value)
            cell.border = border

    end_row = start_row + len(rows)
    end_col = len(headers)

    if rows:
        ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"

        TABLE_COUNTER += 1

        tab = Table(
            displayName=f"EBRIP_Table_{TABLE_COUNTER}",
            ref=ref
        )

        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        ws.add_table(tab)

    return end_row

def autofit(ws):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 45)

def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def pct(value):
    if isinstance(value, (int, float)):
        return value
    return value

# ---------- Clear existing sheets ----------
for ws in wb.worksheets:
    clear_sheet(ws)

# ============================================================
# EXECUTIVE SUMMARY
# ============================================================
ws = wb["Executive Summary"]
write_title(ws, "Enterprise Banking Risk Intelligence Platform — Executive Risk Report")

ws["A3"] = "Purpose"
ws["A3"].font = bold_font
ws["B3"] = (
    "Integrated banking risk analytics covering credit risk, fraud detection, "
    "transaction fraud and AML monitoring."
)

ws["A5"] = "Risk Component"
ws["B5"] = "Model / Engine"
ws["C5"] = "Primary Metric"
ws["D5"] = "Score / Result"
ws["E5"] = "Business Use"

for c in ws[5]:
    c.fill = header_fill
    c.font = header_font
    c.border = border

summary_rows = [
    ["Credit Risk", "Gradient Boosting", "PR-AUC", 0.4091, "Credit default risk assessment"],
    ["Fraud Detection", "Gradient Boosting", "PR-AUC", 0.4077, "Customer / account fraud risk"],
    ["Transaction Fraud", "Random Forest", "PR-AUC", 0.8072, "Transaction-level fraud detection"],
    ["AML Monitoring", "Logistic Regression", "ROC-AUC", 0.8616, "AML transaction monitoring"],
    ["Integrated Risk", "Integrated Risk Engine", "Risk Tier", "LOW", "Cross-domain risk aggregation"],
]

for r, row in enumerate(summary_rows, 6):
    for cidx, value in enumerate(row, 1):
        cell = ws.cell(r, cidx, value)
        cell.border = border

ws["A13"] = "Governance Note"
ws["A13"].font = bold_font
ws["B13"] = (
    "AML validation currently contains only one positive case, therefore AML "
    "threshold metrics are experimental and require a larger representative dataset."
)

# ============================================================
# INTEGRATED RISK
# ============================================================
ws = wb["Integrated Risk"]
write_title(ws, "Integrated Risk Assessment")

integrated = load_json(
    ROOT / "data/staging/integrated_risk/integrated_risk_assessment.json"
)

rows = [
    ["Credit Risk", integrated["credit_risk"]["decision"]],
    ["Fraud Risk", integrated["fraud_risk"]["decision"]],
    ["AML Risk", integrated["aml_risk"]["decision"]],
    ["Overall Risk Score", integrated["integrated_risk"]["overall_risk_score"]],
    ["Risk Tier", integrated["integrated_risk"]["risk_tier"]],
    ["Alert Status", integrated["integrated_risk"]["alert_status"]],
    ["Top Risk Driver", integrated["integrated_risk"]["top_risk_driver"]],
]

write_table(ws, 3, ["Metric", "Value"], rows)

ws["A13"] = "Risk Weights"
ws["A13"].font = bold_font

weights = integrated["risk_weights"]

write_table(
    ws,
    14,
    ["Component", "Weight"],
    [
        ["Credit Risk", weights["credit_risk"]],
        ["Fraud Risk", weights["fraud_risk"]],
        ["AML Risk", weights["aml_risk"]],
    ]
)

# ============================================================
# CREDIT RISK
# ============================================================
ws = wb["Credit Risk"]
write_title(ws, "Credit Risk Model Performance")

credit = load_json(
    ROOT / "data/staging/credit_risk/credit_risk_final_evaluation.json"
)

models = load_json(
    ROOT / "data/staging/credit_risk/credit_risk_model_evaluation.json"
)["models"]

rows = []

for name, metrics in models.items():
    rows.append([
        name.replace("_", " ").title(),
        metrics["accuracy"],
        metrics["precision"],
        metrics["recall"],
        metrics["f1_score"],
        metrics["roc_auc"],
        metrics["pr_auc"],
    ])

write_table(
    ws,
    3,
    ["Model", "Accuracy", "Precision", "Recall", "F1", "ROC-AUC", "PR-AUC"],
    rows
)

ws["A9"] = "Business Threshold"
ws["B9"] = credit["business_threshold"]["threshold"]

ws["A10"] = "Business Recall"
ws["B10"] = credit["business_threshold"]["recall"]

ws["A11"] = "Business Precision"
ws["B11"] = credit["business_threshold"]["precision"]

ws["A12"] = "High Risk"
ws["B12"] = credit["risk_segment_distribution"]["High Risk"]

ws["A13"] = "Very High Risk"
ws["B13"] = credit["risk_segment_distribution"]["Very High Risk"]

# ============================================================
# FRAUD DETECTION
# ============================================================
ws = wb["Fraud Detection"]
write_title(ws, "Fraud Detection Model Performance")

fraud = load_json(
    ROOT / "data/staging/fraud/fraud_final_evaluation.json"
)

fraud_models = load_json(
    ROOT / "data/staging/fraud/fraud_model_evaluation.json"
)["models"]

rows = []

for name, metrics in fraud_models.items():
    rows.append([
        name.replace("_", " ").title(),
        metrics["accuracy"],
        metrics["precision"],
        metrics["recall"],
        metrics["f1_score"],
        metrics["roc_auc"],
        metrics["pr_auc"],
    ])

write_table(
    ws,
    3,
    ["Model", "Accuracy", "Precision", "Recall", "F1", "ROC-AUC", "PR-AUC"],
    rows
)

ws["A9"] = "Selected Model"
ws["B9"] = fraud["model"]["type"]

ws["A10"] = "Best F1 Threshold"
ws["B10"] = fraud["thresholds"]["best_f1"]

ws["A11"] = "Business Threshold"
ws["B11"] = fraud["thresholds"]["business"]

# ============================================================
# TRANSACTION FRAUD
# ============================================================
ws = wb["Transaction Fraud"]
write_title(ws, "Transaction Fraud Detection")

tf = load_json(
    ROOT / "data/staging/transaction_fraud/transaction_fraud_model_evaluation.json"
)

rows = []

for name, metrics in tf["models"].items():
    rows.append([
        name.replace("_", " ").title(),
        metrics["accuracy"],
        metrics["precision"],
        metrics["recall"],
        metrics["f1_score"],
        metrics["roc_auc"],
        metrics["pr_auc"],
    ])

write_table(
    ws,
    3,
    ["Model", "Accuracy", "Precision", "Recall", "F1", "ROC-AUC", "PR-AUC"],
    rows
)

tf_threshold = load_json(
    ROOT / "data/staging/transaction_fraud/transaction_fraud_threshold_metadata.json"
)

ws["A9"] = "Selected Threshold"
ws["B9"] = tf_threshold["selected_threshold"]

ws["A10"] = "Selected Precision"
ws["B10"] = tf_threshold["selected_precision"]

ws["A11"] = "Selected Recall"
ws["B11"] = tf_threshold["selected_recall"]

ws["A12"] = "Selected F1"
ws["B12"] = tf_threshold["selected_f1_score"]

ws["A13"] = "Selection Rule"
ws["B13"] = tf_threshold["selection_reason"]

# ============================================================
# AML MONITORING
# ============================================================
ws = wb["AML Monitoring"]
write_title(ws, "AML Monitoring & Model Governance")

aml = load_json(
    ROOT / "data/staging/aml/aml_reporting/aml_business_report.json"
)

aml_threshold = load_json(
    ROOT / "data/staging/aml/aml_threshold/aml_threshold_metadata.json"
)

rows = [
    ["Model Type", aml["model"]["model_type"]],
    ["Default Threshold", aml["model"]["default_threshold"]],
    ["Best F1 Threshold", aml["model"]["best_f1_threshold"]],
    ["Business Threshold", aml["model"]["business_threshold"]],
    ["Validation Rows", aml_threshold["validation_rows"]],
    ["Positive Cases", aml_threshold["validation_positive_cases"]],
    ["Negative Cases", aml_threshold["validation_negative_cases"]],
    ["ROC-AUC", aml_threshold["roc_auc"]],
    ["PR-AUC", aml_threshold["pr_auc"]],
    ["Governance Status", "Experimental / Requires Larger Dataset"],
]

write_table(ws, 3, ["Metric", "Value"], rows)

# ============================================================
# FRAUD ALERTS
# ============================================================
ws = wb["Fraud Alerts"]
write_title(ws, "Fraud & AML Alert Summary")

alert_files = [
    ROOT / "dashboards/aml/aml_alert_summary.csv",
    ROOT / "dashboards/integrated/integrated_alert_summary.csv",
]

row_number = 3

for file in alert_files:
    if file.exists():
        ws.cell(row_number, 1, file.stem)
        ws.cell(row_number, 1).font = bold_font
        row_number += 1

        with open(file, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            for row in reader:
                for col, value in enumerate(row, 1):
                    ws.cell(row_number, col, value)
                row_number += 1

        row_number += 1

# ============================================================
# MODEL PERFORMANCE
# ============================================================
ws = wb["Model Performance"]
write_title(ws, "Cross-Domain Model Performance")

rows = [
    ["Credit Risk", "Gradient Boosting", 0.871633, 0.409072],
    ["Fraud Detection", "Gradient Boosting", 0.868241, 0.407716],
    ["Transaction Fraud", "Random Forest", 0.972772, 0.807150],
    ["AML Monitoring", "Logistic Regression", 0.861643, 0.000361],
]

write_table(
    ws,
    3,
    ["Risk Domain", "Selected Model", "ROC-AUC", "PR-AUC"],
    rows
)

# ============================================================
# DATA DICTIONARY
# ============================================================
ws = wb["Data Dictionary"]
write_title(ws, "Enterprise Banking Risk Intelligence — Data Dictionary")

dictionary_file = ROOT / "data/raw/credit_risk/Data Dictionary.xls"

rows = [
    ["Credit Risk", "SeriousDlqin2yrs", "Credit default target"],
    ["Fraud Detection", "SeriousDlqin2yrs", "Fraud/default risk target used by current dataset"],
    ["Transaction Fraud", "Class", "Transaction fraud indicator"],
    ["AML", "AML transaction indicators", "AML monitoring features"],
]

write_table(
    ws,
    3,
    ["Domain", "Field / Target", "Description"],
    rows
)

# ---------- Number formatting ----------
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"
    autofit(ws)
    ws.freeze_panes = "A3"

# Save
wb.save(OUTPUT)

print("==========================================")
print("EBRIP EXCEL REPORT GENERATED SUCCESSFULLY")
print("==========================================")
print(f"File: {OUTPUT.resolve()}")
print("Sheets:", ", ".join(wb.sheetnames))

